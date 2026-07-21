#!/usr/bin/env python3
"""
Julkaise ruokalista.html jaettavaksi.

Mitä tekee:
- Päivittää <meta name="julkaistu"> -kentän tämän päivän päivämäärällä
- Kopioi HTML:n julkaisu/ kansioon nimellä index.html (valmis upattavaksi)
- Lisää robots-direktiivin ettei Google löydä sitä turhaan

Käyttö:
    python3 julkaise.py                   # päivitä julkaisu/
    python3 julkaise.py --ei-kopiota      # muuta vain aikaleima ruokalista.html:ssä

Hostaus (valitse yksi):

1) Netlify Drop — helpoin, ei tiliä tarvita kokeiluun:
   - Mene https://app.netlify.com/drop
   - Raahaa julkaisu/-kansio selaimeen
   - Saat URL:n heti (esim. https://ihanajaanimi-12345.netlify.app)
   - Jokaisella päivityksellä: aja tämä skripti uudestaan, raahaa uusi kansio
     → saat saman URL:n jos kirjaudut Netlify-tilille.

2) GitHub Pages — automaattinen uudelleenjulkaisu git push:illa:
   - Luo GitHub-tili ja public-repo nimellä ruokalista
   - cd julkaisu/ && git init && git remote add origin <repo-url>
   - Aktivoi Settings → Pages → Deploy from branch main
   - Jokaisella päivityksellä: git add -A && git commit -m "päivitys" && git push

3) Cloudflare Pages — ilmainen, nopea CDN:
   - https://pages.cloudflare.com → yhdistä GitHub-repo
"""
import argparse
import json
import re
import shutil
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
HTML_PATH = SCRIPT_DIR / "ruokalista.html"
TAPAHTUMAT_HTML_PATH = SCRIPT_DIR / "tapahtumat.html"
RESEPTIT_PATH = SCRIPT_DIR / "reseptit.json"
JULKAISU_DIR = SCRIPT_DIR / "julkaisu"

# Tampere events (Bubster/eventz.today)
TAMPERE_EVENTS_COLLECTION = "634844c32f41a024ee51a234"
TAMPERE_EVENTS_URL = (
    f"https://tapahtumat.tampere.fi/api/collection/{TAMPERE_EVENTS_COLLECTION}/content"
    "?lang=fi&country=FI&mode=event&sort=startDate&count=2000"
)
EVENTZ_IMAGE_BASE = "https://s3.eu-central-1.amazonaws.com/eventz.today.prod/images/"
EVENT_PAGE_BASE = "https://tapahtumat.tampere.fi/fi-FI/page/"

# Takahuhdin päiväkoti, Tampere (Aromi / CGI Saas)
AROMI_BASE = "https://aromimenu.cgisaas.fi/TampereAromieMenus/FI/Default/Tampere/TAKAHUHPK"
AROMI_RESTAURANT_ID = "eb262cfa-9f4a-4616-9b2c-9b9c32c6fa65"
AROMI_DINERGROUP_PAYLOAD = {
    "Id": "64376bd4-d992-427c-af6e-9f20626e19d8",
    "DinerGroupId": "48aeb7f2-7b58-4dca-80e4-a5ce3c4e978e",
    "NutrientGroupId": "8d835066-3834-44c2-abf0-7a83ce375f04",
    "DietGroupId": "b94dc776-277a-4837-a440-4fe9172c3f35",
    "FilterDietGroupId": "7b0bb8d7-ca2b-4b60-9787-e73ee244ab72",
    "DietId": None,
    "ConceptId": "43001f8c-a2e8-478d-9b43-97e7d05aa650",
    "Name": "Päiväkoti", "Code": "Päiväkoti",
    "RestaurantId": AROMI_RESTAURANT_ID,
    "UniqueCode": "Päiväkoti", "IndexNumber": 453, "NameOrCode": "Päiväkoti",
    "WeekDays": ["1", "2", "3", "4", "5"],
    "WeekDay0": False, "WeekDay1": True, "WeekDay2": True, "WeekDay3": True,
    "WeekDay4": True, "WeekDay5": True, "WeekDay6": False,
    "DietType": None, "IsActiveSuitability": False, "SuitabilityDietIds": [],
}


def paivita_aikaleima(html: str, pvm: str) -> str:
    """Päivittää <meta name=\"julkaistu\" content=\"YYYY-MM-DD\">-kentän."""
    pattern = r'(<meta\s+name=["\']julkaistu["\']\s+content=["\'])[^"\']*(["\'])'
    uusi = rf'\g<1>{pvm}\g<2>'
    if not re.search(pattern, html):
        # Lisää meta jos puuttuu — viewportin perään
        html = re.sub(
            r'(<meta\s+name=["\']viewport["\'][^>]*>)',
            rf'\1\n<meta name="julkaistu" content="{pvm}">',
            html,
            count=1,
        )
    else:
        html = re.sub(pattern, uusi, html)
    return html


def lisaa_robots(html: str) -> str:
    """Lisää <meta name=\"robots\" content=\"noindex\"> jos puuttuu."""
    if re.search(r'<meta\s+name=["\']robots["\']', html):
        return html
    return re.sub(
        r'(<meta\s+name=["\']julkaistu["\'][^>]*>)',
        r'\1\n<meta name="robots" content="noindex, nofollow">',
        html,
        count=1,
    )


def hae_paivakodin_valikko(paivia: int = 14) -> dict | None:
    """Hakee Takahuhdin päiväkodin valikon Aromi-API:sta.

    Palauttaa normalisoidun rakenteen tai None jos haku epäonnistuu.
    Muoto:
      {
        "haettu": "2026-04-22T15:10:00",
        "lahde": "Takahuhdin päiväkoti (Aromi)",
        "paivat": {
          "2026-04-22": {
            "lounas": "Pinaattiohukainen, Kermaviilikastike, Perunat, Puolukkahillo",
            "kasvislounas": "Pinaattiohukainen, Kermaviilikastike, Perunat, Puolukkahillo",
            "valipala": "Persikkakiisseli, Täysjyväviipaleet, Edamjuusto, Paprika",
            "aamiainen": "Ohrahiutalepuuro, Täysjyväviipaleet, Kurkku"
          },
          ...
        }
      }
    """
    try:
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        end = (today + timedelta(days=paivia)).replace(hour=23, minute=59, second=59)
        # Aromi haluaa .NET-tyylisen ISO:n ilman aikavyöhykettä
        start_iso = today.strftime("%Y-%m-%dT%H:%M:%S.000Z")
        end_iso = end.strftime("%Y-%m-%dT%H:%M:%S.999Z")
        url = (
            f"{AROMI_BASE}/api/Common/Restaurant/RestaurantMeals"
            f"?Id={AROMI_RESTAURANT_ID}&StartDate={start_iso}&EndDate={end_iso}"
        )
        body = json.dumps(AROMI_DINERGROUP_PAYLOAD, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(
            url,
            data=body,
            method="POST",
            headers={
                "Accept": "application/json",
                "Content-Type": "application/json; charset=utf-8",
                "X-Requested-With": "XMLHttpRequest",
                "User-Agent": "perheen-ruokalista/1.0 (+julkaise.py)",
                "Origin": "https://aromimenu.cgisaas.fi",
                "Referer": f"{AROMI_BASE}/Page/Restaurant",
            },
        )
        with urllib.request.urlopen(req, timeout=20) as resp:
            raw = resp.read().decode("utf-8")
        data = json.loads(raw)
    except Exception as e:
        print(f"⚠ Päiväkodin valikkoa ei saatu: {e}")
        return None

    # Aterianimet → sisäiset avaimet
    meal_key_map = {
        "Aamiainen": "aamiainen",
        "Lounas": "lounas",
        "Lounas A": "lounas",
        "Kasvislounas": "kasvislounas",
        "Kasvislounas A": "kasvislounas",
        "Välipala": "valipala",
        "Välipala A": "valipala",
    }

    def dish_line(dishes):
        names = []
        for d in dishes or []:
            name = (d.get("DishName") or "").strip()
            if not name:
                continue
            # Vältä duplikaatit jos sama ruoka listattu monesti
            if name not in names:
                names.append(name)
        return ", ".join(names)

    paivat = {}
    for day in data if isinstance(data, list) else []:
        pvm_iso = (day.get("Date") or "")[:10]
        if not pvm_iso:
            continue
        paivat[pvm_iso] = {}
        for meal in day.get("Meals") or []:
            key = meal_key_map.get(meal.get("MealName"))
            if not key:
                continue
            line = dish_line(meal.get("Dishes"))
            if line:
                paivat[pvm_iso][key] = line

    if not paivat:
        print("⚠ Päiväkodin valikko oli tyhjä.")
        return None

    return {
        "haettu": datetime.now().isoformat(timespec="seconds"),
        "lahde": "Takahuhdin päiväkoti (Aromi)",
        "paivat": paivat,
    }


def injektoi_paivakoti(html: str, valikko: dict | None) -> str:
    """Upottaa päiväkodin valikon <script id=\"daycare-menu\">-lohkoon."""
    if not valikko:
        return html
    payload = json.dumps(valikko, ensure_ascii=False, separators=(",", ":"))
    pattern = r'(<script\s+type=["\']application/json["\']\s+id=["\']daycare-menu["\']\s*>)([\s\S]*?)(</script>)'
    if re.search(pattern, html):
        return re.sub(pattern, lambda m: m.group(1) + payload + m.group(3), html, count=1)
    # Lisää uusi lohko juuri recipes-data -lohkon jälkeen
    insert = f'\n<script type="application/json" id="daycare-menu">{payload}</script>'
    html2 = re.sub(
        r'(<script\s+type=["\']application/json["\']\s+id=["\']recipes-data["\']\s*>[\s\S]*?</script>)',
        lambda m: m.group(1) + insert,
        html,
        count=1,
    )
    if html2 != html:
        return html2
    # Viimeinen fallback: ennen </body>
    return html.replace("</body>", f"{insert}\n</body>", 1)


def injektoi_reseptit(html: str, reseptit_path: Path) -> str:
    """Upottaa reseptit.json:n <script id=\"recipes-data\">-lohkoon.

    Poistaa yksityiset kentät (last_cooked, palaute) julkaisusta.
    """
    if not reseptit_path.exists():
        return html
    data = json.loads(reseptit_path.read_text(encoding="utf-8"))
    for r in data.get("reseptit", []):
        r.pop("last_cooked", None)
        r.pop("palaute", None)
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    # Korvaa script-lohkon sisältö
    pattern = r'(<script\s+type=["\']application/json["\']\s+id=["\']recipes-data["\']\s*>)([\s\S]*?)(</script>)'
    if not re.search(pattern, html):
        return html  # ei placeholder-lohkoa, ohitetaan
    return re.sub(pattern, lambda m: m.group(1) + payload + m.group(3), html, count=1)


def injektoi_tilaushistoria(html: str, historia_path: Path) -> str:
    """Upottaa historia.json:n <script id="tilaushistoria">-lohkoon.

    Sivuston ostoslistan algoritmi käyttää näitä viimeisimpiä S-kaupat-tilausten
    todellisia hintoja oletusehdotuksena (userPriceFor JS-funktiossa). Lähetetään
    julkaisuun vain `paivitetty`, `hinnat` ja `tilaukset_yhteenveto` — tilausten
    täydet tuotelistat jätetään pois jotta sivu ei paisu turhaan.
    """
    if not historia_path.exists():
        return html
    raw = json.loads(historia_path.read_text(encoding="utf-8"))
    # Rakenna kevennetty payload sivuston tarpeisiin
    tilaukset_yhteenveto = []
    for t in raw.get("tilaukset", []):
        # Älä laske mukaan palvelumaksuja ("Nouto", "Pahvilaatikko")
        n_tuotteita = sum(
            1 for x in t.get("tuotteet", [])
            if x.get("nimi") not in ("Nouto", "Pahvilaatikko")
        )
        tilaukset_yhteenveto.append({
            "id": t.get("id"),
            "pvm": t.get("pvm"),
            "summa": t.get("summa_euroa"),
            "kauppa": t.get("kauppa"),
            "tuotteita": n_tuotteita,
        })
    payload_obj = {
        "paivitetty": raw.get("paivitetty"),
        "hinnat": raw.get("hinnat", {}),
        "tilaukset_yhteenveto": tilaukset_yhteenveto,
    }
    payload = json.dumps(payload_obj, ensure_ascii=False, separators=(",", ":"))
    pattern = r'(<script\s+type=["\']application/json["\']\s+id=["\']tilaushistoria["\']\s*>)([\s\S]*?)(</script>)'
    if not re.search(pattern, html):
        return html
    return re.sub(pattern, lambda m: m.group(1) + payload + m.group(3), html, count=1)


def injektoi_inventaario(html: str, inv_path: Path) -> str:
    """Upottaa kuiva-aine- ja pakastininventaarion <script id="inventaario">-lohkoon.

    Sivuston ostoslista näyttää 🏠/❄-merkin riveille, joiden tuotetta on jo
    kaapissa tai pakastimessa (inventoryMatchFor JS-funktiossa). Vaatii
    openpyxl:n; jos se puuttuu, jätetään olemassa oleva lohko ennalleen.
    """
    if not inv_path.exists():
        return html
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return html
    try:
        wb = openpyxl.load_workbook(inv_path, data_only=True)
    except Exception:
        return html
    items = []
    if "Inventaario" in wb.sheetnames:
        for row in wb["Inventaario"].iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            r = (list(row) + [None] * 9)[:9]
            kat, tuote, merkki, _pkoko, _pkpl, _tila, jaljella, _pe, huom = r
            items.append({
                "varasto": "kuiva", "kategoria": kat, "tuote": str(tuote),
                "merkki": merkki,
                "jaljella_g": jaljella if isinstance(jaljella, (int, float)) else None,
                "huom": huom,
            })
    if "Pakastin" in wb.sheetnames:
        for row in wb["Pakastin"].iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            r = (list(row) + [None] * 8)[:8]
            kat, tuote, merkki, maara, yks, _pvm, _pe, huom = r
            g = maara if (isinstance(maara, (int, float)) and str(yks).lower() == "g") else None
            items.append({
                "varasto": "pakastin", "kategoria": kat, "tuote": str(tuote),
                "merkki": merkki, "jaljella_g": g, "huom": huom,
            })
    payload_obj = {
        "paivitetty": date.today().isoformat(),
        "lahde": str(inv_path.name),
        "tuotteet": items,
    }
    payload = json.dumps(payload_obj, ensure_ascii=False, separators=(",", ":"))
    pattern = r'(<script\s+type=["\']application/json["\']\s+id=["\']inventaario["\']\s*>)([\s\S]*?)(</script>)'
    if not re.search(pattern, html):
        return html
    return re.sub(pattern, lambda m: m.group(1) + payload + m.group(3), html, count=1)


def _slugify_name(name: str) -> str:
    """Tampere-sivuston URL-slug: pienet kirjaimet, ääkköset, välilyönnit viivoiksi."""
    if not name:
        return ""
    import urllib.parse
    s = name.strip().lower()
    # Korvaa tavalliset välimerkit välilyönnellä, sitten välilyönnit viivoiksi
    s = re.sub(r"[!?.,:;/\\\"'()]", "", s)
    s = re.sub(r"\s+", "-", s)
    return urllib.parse.quote(s, safe="-äöåÄÖÅ")


def _hae_tampere_tapahtumat_raw(aikakatkaisu: int = 30) -> list | None:
    """Hakee Tampereen tapahtumat Bubster-APIsta."""
    try:
        req = urllib.request.Request(
            TAMPERE_EVENTS_URL,
            headers={
                "Accept": "application/json",
                "User-Agent": "perheen-ruokalista/1.0 (+julkaise.py)",
            },
        )
        with urllib.request.urlopen(req, timeout=aikakatkaisu) as resp:
            raw = resp.read().decode("utf-8")
        data = json.loads(raw)
    except Exception as e:
        print(f"⚠ Tampereen tapahtumia ei saatu: {e}")
        return None
    pages = data.get("pages")
    if not isinstance(pages, list):
        print("⚠ Odottamaton tapahtumadata — ei 'pages'-listaa")
        return None
    return pages


def _normalisoi_tapahtuma(p: dict, nyt: datetime, loppu: datetime) -> dict | None:
    """Normalisoi yhden tapahtuman sivun käyttöön. Palauttaa None jos ei sovi aikaikkunaan."""
    ev = p.get("event") or {}
    # Etsi ensimmäinen tuleva (tai käynnissä oleva) alkamisaika
    alkaa_iso = None
    paattyy_iso = None
    dates = ev.get("dates") or []
    for d in dates:
        start = d.get("start")
        if not start:
            continue
        try:
            dt = datetime.fromisoformat(start.replace("Z", "+00:00"))
        except ValueError:
            continue
        if dt >= nyt - timedelta(hours=12) and dt <= loppu:
            alkaa_iso = start
            paattyy_iso = d.get("end")
            break
    # Fallback: event.start
    if not alkaa_iso and ev.get("start"):
        try:
            dt = datetime.fromisoformat(ev["start"].replace("Z", "+00:00"))
            if dt >= nyt - timedelta(hours=12) and dt <= loppu:
                alkaa_iso = ev["start"]
                paattyy_iso = ev.get("end")
        except ValueError:
            pass
    if not alkaa_iso:
        return None

    # Paikka
    paikka = ""
    for loc in p.get("locations") or []:
        addr = (loc or {}).get("address")
        if addr:
            paikka = addr
            break

    # Kuva
    img_id = p.get("imageList") or p.get("imageDesktop") or p.get("imageMobile") or ""
    kuva = (EVENTZ_IMAGE_BASE + img_id) if img_id else ""

    # Linkki tapahtumasivulle
    pid = p.get("_id") or ""
    nimi = (p.get("name") or "").strip()
    linkki = ""
    if pid and nimi:
        linkki = EVENT_PAGE_BASE + pid + "/" + _slugify_name(nimi)

    # Tagit: hashtagit ensin, sitten järjestäjä tunnisteena
    tagit = []
    for h in p.get("hashtags") or []:
        if h and h not in tagit:
            tagit.append(h)

    return {
        "id": pid,
        "nimi": nimi,
        "kuvaus": (p.get("descriptionShort") or "").strip(),
        "alkaa": alkaa_iso,
        "paattyy": paattyy_iso,
        "paikka": paikka,
        "jarjestaja": (p.get("ownerName") or "").strip(),
        "kohderyhmat": p.get("ages") or [],
        "kategoriat": p.get("globalContentCategories") or [],
        "tagit": tagit,
        "kuva": kuva,
        "linkki": linkki,
    }


def _sopiiko_sivulle(t: dict) -> bool:
    """True jos tapahtuma sopii perhe- tai konservatorio-välilehdelle."""
    # Konservatorio
    jarj = (t.get("jarjestaja") or "").lower()
    if "konservator" in jarj:
        return True
    # Perhe: age-1 tai age-2, tai "kids and family" -kategoria
    ages = t.get("kohderyhmat") or []
    if "age-1" in ages or "age-2" in ages:
        return True
    kat = [str(c).lower() for c in (t.get("kategoriat") or [])]
    for c in kat:
        if "kids and family" in c or "lapsi" in c:
            return True
    return False


def hae_tampere_tapahtumat(kuukaudet: int = 3) -> dict | None:
    """Hakee ja suodattaa Tampereen tapahtumat.

    Karsii julkaisuvaiheessa vain ne jotka sivu näyttää (perhe + konservatorio),
    jotta HTML pysyy pienenä mobiilia varten.

    Palauttaa rakenteen:
      { "haettu": "2026-04-22T19:00:00", "tapahtumat": [...] }
    """
    pages = _hae_tampere_tapahtumat_raw()
    if pages is None:
        return None
    from datetime import timezone
    nyt = datetime.now(timezone.utc)
    loppu = nyt + timedelta(days=kuukaudet * 31)
    normalisoidut = []
    for p in pages:
        row = _normalisoi_tapahtuma(p, nyt, loppu)
        if not row:
            continue
        if not _sopiiko_sivulle(row):
            continue
        normalisoidut.append(row)

    # Lajittele alkamisajan mukaan
    def _key(r):
        try:
            return datetime.fromisoformat((r.get("alkaa") or "").replace("Z", "+00:00"))
        except Exception:
            return datetime.max.replace(tzinfo=timezone.utc)
    normalisoidut.sort(key=_key)

    return {
        "haettu": datetime.now().isoformat(timespec="seconds"),
        "tapahtumat": normalisoidut,
    }


def injektoi_tapahtumat(html: str, tapahtumat: dict | None) -> str:
    """Upottaa Tampereen tapahtumat <script id=\"events-data\">-lohkoon."""
    if not tapahtumat:
        return html
    payload = json.dumps(tapahtumat, ensure_ascii=False, separators=(",", ":"))
    pattern = r'(<script\s+type=["\']application/json["\']\s+id=["\']events-data["\']\s*>)([\s\S]*?)(</script>)'
    if re.search(pattern, html):
        return re.sub(pattern, lambda m: m.group(1) + payload + m.group(3), html, count=1)
    return html


def _html_escape(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _kerää_näkyvät_reseptinimet(html: str) -> set[str]:
    """Kerää vk1 + vk2 päivällisillä jo näkyvät reseptinimet (vältettäväksi vk3:lla)."""
    nimet = set()
    for m in re.finditer(
        r'data-k="w[12]-d-(?:ma|ti|ke|to|pe|la|su)"[^>]*>\s*<strong>([^<]+)</strong>',
        html,
    ):
        nimet.add(m.group(1).strip())
    return nimet


def injektoi_viikkojen_paivalliset(
    html: str,
    viikot: tuple[str, ...] = ("w1", "w2", "w3"),
    viikon_alkupvmt: dict | None = None,
) -> tuple[str, dict[str, list[str]]]:
    """Generoi useamman viikon päivälliset ja injektoi HTML:n DEFAULT-soluihin.

    Korjaa rotaatiobugin: kun julkaise.py ajetaan, kaikki näkyvät viikot
    päivitetään, jotta DEFAULT_CELLS edustaa aina nykyhetken 3-viikon ikkunaa
    (ennen tätä vain vk 3 päivittyi → vk 1 ja vk 2 jäivät edellisten ajojen
    arvoilla, jolloin viikon vaihtuessa "tämä viikko" -välilehti näytti
    edellisen viikon reseptit).

    Käyttää ehdota_viikko.py:n logiikkaa: priorisoi reseptejä joita ei ole
    keitetty pitkään aikaan, ei toistoja 15 arkipäivällisen joukossa.
    Käyttäjän localStorage (state.meals[YYYY-MM-DD]) voittaa renderöinnissä,
    eli käsin asetetut päivät säilyvät.

    Palauttaa (uusi_html, {viikko: [nimet]}).
    """
    try:
        from ehdota_viikko import suodata, valitse, lataa
    except Exception as e:
        print(f"⚠ Viikkoehdotuksia ei voitu generoida: {e}")
        return html, {}

    try:
        data = lataa()
        reseptit = data.get("reseptit", [])
    except Exception as e:
        print(f"⚠ Reseptien lataus epäonnistui: {e}")
        return html, {}

    paivalliset = [r for r in reseptit if r.get("kategoria") == "päivällinen"]
    # Sanity check: jos päivällisreseptejä on alle minimimäärän, jokin on rikki
    if len(paivalliset) < 5:
        from collections import Counter
        kategoriat = Counter(r.get("kategoria", "?") for r in reseptit)
        print(
            f"⚠ KRIITTINEN: vain {len(paivalliset)} päivällisreseptiä löytyi "
            f"({len(reseptit)} reseptiä yhteensä). Kategoriat: {dict(kategoriat)}. "
            f"Jos kategoriat ovat 'pÃ¤ivÃ¤llinen' jne., reseptit.json on "
            f"mojibake-vikainen. HTML-soluja EI päivitetä jotta vanhat sisällöt säilyvät."
        )
        return html, {}

    # PLANNED-OVERRIDE: reseptit.json voi sisältää 'planned' dictin
    # {"YYYY-MM-DD": "recipe_id"} joka pakottaa kyseisen reseptin
    # tiettyyn päivään (lock-suunnitelma). Estää algoritmia korvaamasta sitä.
    # Sallitaan any-kategorian resepti planned-päiviin (esim. lasagne viikonloppuun).
    planned = data.get("planned", {}) if isinstance(data.get("planned"), dict) else {}
    by_id_kaikki = {r.get("id"): r for r in reseptit}

    arki_pool = suodata(paivalliset, viikonloppu=False)
    # Vk-pool: viikonloppu sallii pidempiä reseptejä (sunnuntain ehdotuksiin)
    vk_pool = suodata(paivalliset, viikonloppu=True)

    # Kerää KAIKKI planned-päivien id:t (myös niiden viikkojen joiden alkupvm
    # ei ole annettu, esim. nykyisen viikon planned-päivät kun käsittelyssä on
    # vain ensi vk + +2 vk). Estää että w2:n planned-resepti valitaan myös
    # w3:lle — bug joka aiheutti vk 19 ja vk 20 saman sisällön.
    # Ulkopuoliset planned-id:t lisätään pool-suodatukseen, eivät planned_in_window-
    # mappaukseen (jälkimmäistä käytetään vain "tämän ikkunan" lock-suunnitelmaan).
    paivat = ["ma", "ti", "ke", "to", "pe", "la", "su"]
    planned_ids_in_window: set = set()
    if viikon_alkupvmt:
        for viikko in viikot:
            alku = viikon_alkupvmt.get(viikko)
            if not alku:
                continue
            for i in range(7):
                target_iso = (alku + timedelta(days=i)).isoformat()
                if target_iso in planned:
                    planned_ids_in_window.add(planned[target_iso])

    # Kaikki planned id:t (myös ikkunan ulkopuoliset) — suodatetaan pool:sta jotta
    # ne eivät tule duplikaateiksi. valid-id:t = ne jotka oikeasti löytyvät.
    # KORJAUS 15.7.2026: poissuljetaan vain tuoreet/tulevat planned-varaukset
    # (viim. 3 vk, vastaa "ei toistoja 15 arkipäivällisen joukossa" -sääntöä).
    # Koko planned-historian poissulkeminen kuristi ehdotuspoolin tyhjiin,
    # jolloin w4 jäi tyhjäksi ja auto-updaten sanity-check kaatui (4.–15.7.2026).
    raja_planned = (date.today() - timedelta(days=21)).isoformat()
    planned_ids_kaikki: set = {
        rid for pvm, rid in planned.items()
        if rid in by_id_kaikki and pvm >= raja_planned
    }

    # Algoritmi: 5 arkipäivää (ma-pe) + 1 sunnuntai per viikko = 6 reseptiä/vk.
    # La = tyhjä (viikon ylijäämät), planned-päivät korvataan myöhemmin.
    raja = date.today() - timedelta(days=7)
    tuoreet_arki = [
        r for r in arki_pool
        if (not r.get("last_cooked")
            or datetime.fromisoformat(r["last_cooked"]).date() < raja)
        and r.get("id") not in planned_ids_kaikki
    ]
    yhteensa_arki = 5 * len(viikot)
    arki_valinta = valitse(tuoreet_arki, yhteensa_arki)
    if len(arki_valinta) < yhteensa_arki:
        valitut_idt = {r["id"] for r in arki_valinta} | planned_ids_kaikki
        loput = [r for r in arki_pool if r["id"] not in valitut_idt]
        taydennys = valitse(loput, yhteensa_arki - len(arki_valinta))
        if taydennys:
            print(f"  (arki-pool täydennetty {len(taydennys)} reseptillä)")
        arki_valinta = arki_valinta + taydennys

    # Sunnuntain valinta erikseen — käytetään vk-pool jotta pidemmät reseptit
    # (esim. uunijuurekset, lasagne) saavat priorisointia. Vältetään duplikaatit
    # arki-valinnan JA kaikkien planned-päivien kanssa.
    valitut_idt_kaikki = {r["id"] for r in arki_valinta} | planned_ids_kaikki
    yhteensa_su = len(viikot)
    su_valinta = valitse(vk_pool, yhteensa_su, vältä_ideja=list(valitut_idt_kaikki))

    tulos: dict[str, list[str]] = {}

    for vk_idx, viikko in enumerate(viikot):
        nimet: list[str] = []
        viikon_alku = viikon_alkupvmt.get(viikko) if viikon_alkupvmt else None

        for i, dk in enumerate(paivat):
            target_iso = (viikon_alku + timedelta(days=i)).isoformat() if viikon_alku else None

            new_inner = ""
            r = None

            # 1) PLANNED-override (lock) — pakotettu resepti tiettyyn päivään
            if target_iso and target_iso in planned:
                planned_id = planned[target_iso]
                r = by_id_kaikki.get(planned_id)
                if not r:
                    print(f"⚠ planned[{target_iso}] = {planned_id!r} — reseptiä ei löydy")

            # 2) Algoritmin valinta — ma-pe + su (la jää tyhjäksi)
            elif dk in ("ma", "ti", "ke", "to", "pe"):
                slot_idx = vk_idx * 5 + i
                if slot_idx < len(arki_valinta):
                    r = arki_valinta[slot_idx]
            elif dk == "su":
                if vk_idx < len(su_valinta):
                    r = su_valinta[vk_idx]
            # dk == "la" → r jää None → solu tyhjäksi (= ylijäämät)

            if r:
                nimi = (r.get("nimi") or "").strip()
                if nimi:
                    new_inner = f'<strong>{_html_escape(nimi)}</strong>'
                    nimet.append(nimi)

            pattern = re.compile(
                rf'(<div\s+class="cell"\s+data-k="{viikko}-d-{dk}"[^>]*>)([\s\S]*?)(</div>)'
            )
            replaced = [False]

            def repl(m):
                replaced[0] = True
                return m.group(1) + new_inner + m.group(3)

            html = pattern.sub(repl, html, count=1)
            if not replaced[0]:
                print(f"⚠ Päivällissolu {viikko}-d-{dk} ei löytynyt — sisältöä ei muutettu")

        tulos[viikko] = nimet

    return html, tulos


def injektoi_viikon_toteuma(
    html: str, viikko_id: str = "w1", viikon_alkupvm: date | None = None
) -> tuple[str, list[str]]:
    """Täyttää viikon päivällissolut TOTEUMAN perusteella — eli last_cooked-päivämäärää
    vasten. Tällä viikolla eilen syöty Halloumi näkyy eilen, ei minkään uuden ehdotuksen
    sijaan.

    Jos jollekin päivälle ei löydy reseptiä jonka last_cooked == sen päivän pvm, solu
    jätetään tyhjäksi (käyttäjä voi lisätä omat last_cookedin ulkopuoliset merkinnät
    klikkaamalla solua selaimessa, tai ajaa `python3 ehdota_viikko.py --vahvista YYYY-MM-DD`).

    Palauttaa (uusi_html, [nimet päivien järjestyksessä, tyhjät jätetään pois]).
    """
    try:
        from ehdota_viikko import lataa
        data = lataa()
        reseptit = data.get("reseptit", [])
    except Exception as e:
        print(f"⚠ Toteuman lataus epäonnistui: {e}")
        return html, []

    if viikon_alkupvm is None:
        # Tämän viikon maanantai (Python: weekday() 0=Ma, 6=Su)
        today = date.today()
        viikon_alkupvm = today - timedelta(days=today.weekday())

    # Indeksoi reseptit last_cooked-päivän mukaan
    by_date: dict[str, dict] = {}
    for r in reseptit:
        lc = r.get("last_cooked")
        if lc and lc not in by_date:
            by_date[lc] = r

    # PLANNED-override (lock-suunnitelma): jos last_cooked-toteumaa ei ole,
    # tarkistetaan myös onko tälle päivälle pakotettu resepti planned-dictissä.
    planned = data.get("planned", {}) if isinstance(data.get("planned"), dict) else {}
    by_id = {r.get("id"): r for r in reseptit}

    paivat = ["ma", "ti", "ke", "to", "pe", "la", "su"]
    nimet: list[str] = []

    for i, dk in enumerate(paivat):
        target = (viikon_alkupvm + timedelta(days=i)).isoformat()
        # 1) Toteuma (last_cooked) ensisijaisena
        match = by_date.get(target)
        # 2) Planned overridena jos toteumaa ei ole
        if not match and target in planned:
            match = by_id.get(planned[target])
        if match:
            nimi = (match.get("nimi") or "").strip()
            if nimi:
                new_inner = f'<strong>{_html_escape(nimi)}</strong>'
                nimet.append(nimi)
            else:
                new_inner = ""
        else:
            new_inner = ""

        pattern = re.compile(
            rf'(<div\s+class="cell"\s+data-k="{viikko_id}-d-{dk}"[^>]*>)([\s\S]*?)(</div>)'
        )
        replaced = [False]

        def repl(m):
            replaced[0] = True
            return m.group(1) + new_inner + m.group(3)

        html = pattern.sub(repl, html, count=1)
        if not replaced[0]:
            print(f"⚠ Toteumasolu {viikko_id}-d-{dk} ei löytynyt — sisältöä ei muutettu")

    return html, nimet


# Vanhan nimen yhteensopivuusalias — vk1 päivittyy nyt TOTEUMAN mukaan
# (last_cooked-päivämäärillä), vk2 ja vk3 saavat ehdotukset.
# Näin "tämä viikko" -näkymä näyttää aidot syömämme ateriat eikä uudelleengeneroituja
# ehdotuksia, kun taas vk2/vk3 ovat suunnittelu­näkymiä.
def injektoi_viikon3_paivalliset(html: str) -> tuple[str, list[str]]:
    html, w1_nimet = injektoi_viikon_toteuma(html, "w1")
    if w1_nimet:
        print(f"✓ Vk 1 toteuma (last_cooked): {', '.join(w1_nimet)}")
    else:
        print("  Vk 1 toteuma tyhjä — ei kokattuja reseptejä tällä viikolla "
              "(käytä `python3 ehdota_viikko.py --vahvista YYYY-MM-DD` merkitäksesi)")
    html, tulos = injektoi_viikkojen_paivalliset(html, ("w2", "w3"))
    return html, tulos.get("w3", [])


def injektoi_kesaloma_lounaat(
    html: str,
    viikkojen_ankkurit: dict[str, date],
) -> tuple[str, dict[str, list[str]]]:
    """Kesäloman aikana päivän D lounas-soluun täytetään päivän D-1 päivällinen "tähteinä".

    Lukee reseptit.json:n kesaloma-kentästä jakson (alku, loppu) ja käy kaikki annettujen
    viikkojen päivät läpi. Jos päivä D on jaksolla, etsii sivun jo täytetyn päivällissolun
    (data-k="wN-d-pp") päivälle D-1 ja kopioi reseptin lounas-soluun muodossa
    "<strong>Tähteet</strong><br>RESEPTI".

    Tämä funktio pitää kutsua SEN JÄLKEEN kun kaikki päivälliset on jo täytetty
    (injektoi_viikon_toteuma + injektoi_viikkojen_paivalliset).

    Toimii viikkorajojen yli: ma:n lounas = edellisen viikon su:n päivällinen.

    Palauttaa (uusi_html, {viikko: [täytetyt lounaspäivät]}).
    """
    try:
        with open(RESEPTIT_PATH, encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print(f"⚠ Reseptien lataus kesäloma-lounaita varten epäonnistui: {e}")
        return html, {}

    kesa = data.get("kesaloma")
    if not isinstance(kesa, dict) or "alku" not in kesa or "loppu" not in kesa:
        return html, {}

    try:
        kesa_alku = date.fromisoformat(kesa["alku"])
        kesa_loppu = date.fromisoformat(kesa["loppu"])
    except Exception as e:
        print(f"⚠ Kesaloma-kentän alku/loppu virheellinen: {e}")
        return html, {}

    paivat = ["ma", "ti", "ke", "to", "pe", "la", "su"]

    # Käänteismappays: pvm → (viikko_id, dk) — jotta D-1 päivä löytyy nopeasti
    pvm_to_solu: dict[date, tuple[str, str]] = {}
    for vk_id, ankkuri in viikkojen_ankkurit.items():
        for i, dk in enumerate(paivat):
            pvm_to_solu[ankkuri + timedelta(days=i)] = (vk_id, dk)

    def poimi_paivallinen(vk_id: str, dk: str) -> str | None:
        """Poimii <strong>...</strong>-sisällön päivällissolusta. Palauttaa None
        jos solua ei löydy tai se on tyhjä."""
        pat = re.compile(
            rf'<div\s+class="cell"\s+data-k="{vk_id}-d-{dk}"[^>]*>([\s\S]*?)</div>'
        )
        m = pat.search(html)
        if not m:
            return None
        m2 = re.search(r'<strong>([^<]*)</strong>', m.group(1))
        if not m2:
            return None
        return m2.group(1).strip() or None

    tulos: dict[str, list[str]] = {}

    for vk_id, ankkuri in viikkojen_ankkurit.items():
        nimet: list[str] = []
        for i, dk in enumerate(paivat):
            pvm = ankkuri + timedelta(days=i)
            on_kesaloma = kesa_alku <= pvm <= kesa_loppu

            # KORJAUS 21.7.2026: solu kirjoitetaan AINA (myös tyhjäksi), jotta
            # vanhojen buildien jäänteet eivät jää näkyviin. Säännöt:
            # - kesäloma tai viikonloppu (la/su): tähteet edellisen päivän
            #   päivällisestä, jos sellainen on
            # - kouluviikon ma–pe (kesäloman ulkopuolella): tyhjä
            # - su ilman tähteitä: "Viikon ylijäämät" -oletus
            paivallinen_nimi = None
            if on_kesaloma or dk in ("la", "su"):
                edellinen_solu = pvm_to_solu.get(pvm - timedelta(days=1))
                if edellinen_solu:
                    paivallinen_nimi = poimi_paivallinen(*edellinen_solu)

            if paivallinen_nimi:
                new_inner = (
                    f'<strong>Tähteet</strong><br>{_html_escape(paivallinen_nimi)}'
                )
            elif dk == "su":
                new_inner = '<strong>Viikon ylijäämät</strong>'
                paivallinen_nimi = "Viikon ylijäämät"
            else:
                new_inner = ""

            pattern = re.compile(
                rf'(<div\s+class="cell"\s+data-k="{vk_id}-l-{dk}"[^>]*>)'
                rf'([\s\S]*?)(</div>)'
            )
            replaced = [False]

            def repl(m):
                replaced[0] = True
                return m.group(1) + new_inner + m.group(3)

            html = pattern.sub(repl, html, count=1)
            if replaced[0] and paivallinen_nimi:
                nimet.append(f"{dk}={paivallinen_nimi}")
            else:
                print(
                    f"⚠ Lounas-solu {vk_id}-l-{dk} ei löytynyt — "
                    f"sisältöä ei muutettu"
                )

        tulos[vk_id] = nimet

    return html, tulos


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ei-kopiota", action="store_true",
                    help="Päivitä vain aikaleima, älä luo julkaisu-kansiota")
    ap.add_argument("--pvm", default=None,
                    help="Aseta tiettty päivämäärä (YYYY-MM-DD), oletus tänään")
    ap.add_argument("--skip-vk3", action="store_true",
                    help="Älä päivitä vk 3 -päivällisehdotuksia")
    args = ap.parse_args()

    pvm = args.pvm or date.today().isoformat()
    html = HTML_PATH.read_text(encoding="utf-8")
    html_paivitetty = paivita_aikaleima(html, pvm)
    html_paivitetty = lisaa_robots(html_paivitetty)
    html_paivitetty = injektoi_reseptit(html_paivitetty, RESEPTIT_PATH)

    valikko = hae_paivakodin_valikko(paivia=14)
    html_paivitetty = injektoi_paivakoti(html_paivitetty, valikko)

    # Päivällisehdotukset kaikille kolmelle viikolle (default-arvot HTML:ään;
    # käyttäjän localStorage voittaa renderöinnissä).
    if not args.skip_vk3:
        html_paivitetty, vk_nimet = injektoi_viikkojen_paivalliset(
            html_paivitetty, ("w1", "w2", "w3")
        )
        for vk, nimet in vk_nimet.items():
            if nimet:
                print(f"✓ {vk} päivälliset ehdotettu: {', '.join(nimet)}")

    HTML_PATH.write_text(html_paivitetty, encoding="utf-8")
    print(f"✓ Päivitetty aikaleima: {pvm} → ruokalista.html")
    if RESEPTIT_PATH.exists():
        print(f"✓ Reseptit upotettu → recipes-data -lohko")
    if valikko:
        print(f"✓ Päiväkodin valikko upotettu ({len(valikko['paivat'])} päivää)")

    # Tapahtumasivu
    tapahtumat_data = None
    if TAPAHTUMAT_HTML_PATH.exists():
        tap_html = TAPAHTUMAT_HTML_PATH.read_text(encoding="utf-8")
        tap_html = paivita_aikaleima(tap_html, pvm)
        tap_html = lisaa_robots(tap_html)
        tapahtumat_data = hae_tampere_tapahtumat(kuukaudet=3)
        tap_html = injektoi_tapahtumat(tap_html, tapahtumat_data)
        TAPAHTUMAT_HTML_PATH.write_text(tap_html, encoding="utf-8")
        if tapahtumat_data:
            print(f"✓ Tampereen tapahtumat upotettu ({len(tapahtumat_data['tapahtumat'])} kpl)")

    if args.ei_kopiota:
        return

    JULKAISU_DIR.mkdir(exist_ok=True)
    kohde = JULKAISU_DIR / "index.html"
    shutil.copy2(HTML_PATH, kohde)
    print(f"✓ Kopio → {kohde}")
    if TAPAHTUMAT_HTML_PATH.exists():
        kohde_tap = JULKAISU_DIR / "tapahtumat.html"
        shutil.copy2(TAPAHTUMAT_HTML_PATH, kohde_tap)
        print(f"✓ Kopio → {kohde_tap}")

    # Päivitä vain viimeinen rivi README:ssä; säilytä olemassa oleva ohje
    readme = JULKAISU_DIR / "README.txt"
    if readme.exists():
        sisalto = readme.read_text(encoding="utf-8")
        sisalto = re.sub(
            r"Viimeisin julkaisu: \d{4}-\d{2}-\d{2}",
            f"Viimeisin julkaisu: {pvm}",
            sisalto,
        )
        readme.write_text(sisalto, encoding="utf-8")
        print(f"✓ README:n aikaleima päivitetty → {readme}")
    else:
        readme.write_text(
            f"Tämä kansio on valmis upattavaksi GitHub Pagesiin, Netlifyyn tai Cloudflare Pagesiin.\n"
            f"Viimeisin julkaisu: {pvm}\n",
            encoding="utf-8",
        )
        print(f"✓ README luotu → {readme}")
    print()
    print("Seuraavaksi:")
    print(f"  → Seuraa README.txt:n ohjeita julkaisuun GitHub Pagesilla")
    print(f"  → Tai raahaa kansio https://app.netlify.com/drop (nopein testi)")


if __name__ == "__main__":
    main()
